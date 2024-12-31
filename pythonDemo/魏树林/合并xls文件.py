import pandas as pd
import openpyxl 
from openpyxl.styles import Font, colors
from openpyxl.styles import PatternFill
import os


#实现批量读取--被查询必对的文件夹
dir_path=r'C:\Users\17653\Desktop\反洗钱1558户\历史数据查询266911'
#获取指定路径下的文件名
file_names=os.listdir(dir_path)
#获取指定路径下文件名绝对路径
file_paths = [os.path.join(dir_path, file) for file in file_names]
#设定单元格样式
font = Font(name='Arial', size=12, bold=True, italic=False, color=colors.BLUE)

#需要查询的名单表格
source_path=r'C:\Users\17653\Desktop\反洗钱1558户\人行下发解除惩戒名单(1).xlsx'
source_book = openpyxl.load_workbook(source_path,'r')




for item in file_paths:
        #读取 Excel 文件中的所有 Sheet 共计105个表格
        book = openpyxl.load_workbook(item)      
        sheet=book.create_sheet('copysheet')   

        for row in source_book['给银行'].iter_rows():
                #注意rowlist的位置，如写在外面，会造成把复制过来的sheet 按一行排列
                row_list=[]
                for cell in row:
                         row_list.append(cell.value)
                sheet.append(row_list)   

        book.save(item)
        book.close()
        print('复制成功!')

print('执行完成!')
                        

