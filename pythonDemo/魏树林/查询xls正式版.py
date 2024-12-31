import pandas as pd
import openpyxl 
from openpyxl.styles import Font, colors
from openpyxl.styles import PatternFill
import os


#实现批量读取
dir_path=r'C:\Users\17653\Desktop\反洗钱1558户\历史数据查询266911'
#获取指定路径下的文件名
file_names=os.listdir(dir_path)
#获取指定路径下文件名绝对路径
file_paths = [os.path.join(dir_path, file) for file in file_names]
#设定单元格样式
font = Font(name='Arial', size=12, bold=True, italic=False, color=colors.BLUE)

for item in file_paths:
        #读取 Excel 文件中的所有 Sheet 共计105个表格
        book = openpyxl.load_workbook(item)
        #设置单元格颜色
        #获取sheet1表格
        s=book['copysheet']
        #根据sheet1的第三列单元格信息-身份证取查询对应的表格sheet名
        for  i in range(2,1560):
            if  s.cell(i,3).value in book.sheetnames:
                #打开对应sheet1表格对应的sheet命名表格
                sheet=book[s.cell(i,3).value]
                if  sheet["A2"].value is None:#表示只有一行的被删除
                    s.cell(i,4).value="未查询到该账号信息"
                else:
                        s.cell(i,4).value="未查询到该账号信息"
                        s.cell(i,4).value=sheet["A7"].value
                        s.cell(i,5).value=sheet["E7"].value
                        s.cell(i,6).value=sheet["H7"].value
                        s.cell(i,7).value=sheet["G7"].value 

                        if sheet["A8"].value is None:#\表示只有一行的被删除
                            s.cell(i,8).value="只有一个账户"
                        else:
                            s.cell(i,8).value=sheet["A8"].value
                            s.cell(i,9).value=sheet["E8"].value
                            s.cell(i,10).value=sheet["H8"].value
                            s.cell(i,11).value=sheet["G8"].value 

                            if sheet["A9"].value is None:#\表示只有一行的被删除
                                s.cell(i,12).value="仅有两个账户"
                            else:    
                                s.cell(i,12).value=sheet["A9"].value
                                s.cell(i,13).value=sheet["E9"].value
                                s.cell(i,14).value=sheet["H9"].value
                                s.cell(i,15).value=sheet["G9"].value 

                                if sheet["A10"].value is None:#\表示只有一行的被删除
                                    s.cell(i,16).value="仅有三个账户"
                                else:    
                                    s.cell(i,16).value=sheet["A10"].value
                                    s.cell(i,17).value=sheet["E10"].value
                                    s.cell(i,18).value=sheet["H10"].value
                                    s.cell(i,19).value=sheet["G10"].value 
            else:

                #不包含在此xls表格内的标记蓝色
                s.cell(i,3).fill=PatternFill("solid", fgColor="3CB371")
                s.cell(i,4).value='缺少该身份证sheet表格'
        #获取文件的扩展名
        resext=os.path.basename(item)  
        #扩展名分割文件名绝对路径    
        res,_=os.path.splitext(resext)
        #拼接字符串组成新的文件名
        result=res+'查询结果.xlsx'
        #注意以上变化save后才执行成功！
        book.save(result)
        book.close()
        print("写入成功！")

print('执行完毕')



        
