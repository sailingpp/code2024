from openpyxl import Workbook
import pandas as pd
from Pile import *

list_kc1=[
     {"layerId": 1, "layerName": "粘土", "startHeight": 19,  "endHeight": 18, "qsi": 20, "qua": 20} ,
     {"layerId": 2, "layerName": "粉土", "startHeight": 18,  "endHeight": 15, "qsi": 30, "qua": 20} ,
     {"layerId": 3, "layerName": "砂土", "startHeight": 15,  "endHeight": 13, "qsi": 50, "qua": 20} ,
     {"layerId": 4, "layerName": "岩土", "startHeight": 13,  "endHeight": 12, "qsi": 60, "qua": 200},
     {"layerId": 5, "layerName": "基岩", "startHeight": 12,  "endHeight": 10, "qsi": 70, "qua": 200},
   ]
list_kc2=[
     {"layerId": 1, "layerName": "粘土", "startHeight": 19,  "endHeight": 18, "qsi": 20, "qua": 20} ,
     {"layerId": 2, "layerName": "粉土", "startHeight": 18,  "endHeight": 14, "qsi": 30, "qua": 20} ,
     {"layerId": 3, "layerName": "砂土", "startHeight": 14,  "endHeight": 13, "qsi": 50, "qua": 20} ,
     {"layerId": 4, "layerName": "岩土", "startHeight": 13,  "endHeight": 12, "qsi": 60, "qua": 200},
     {"layerId": 5, "layerName": "基岩", "startHeight": 12,  "endHeight": 9, "qsi": 70, "qua": 200},
   ]

dict_soildata={'kc1':list_kc1,'kc2':list_kc2}


df=pd.DataFrame(dict_soildata)

pile=Pile(0.8)
workbook=Workbook()
sheet=workbook.active
sheet.cell(1,1,'直径')
sheet.cell(2,1,pile.diameter)
sheet.cell(3,1,'周长')
sheet.cell(4,1,pile.cir)
sheet.cell(5,1,'面积')
sheet.cell(6,1,pile.area)

sr=8
j=0
for k in dict_soildata:
    for  x in range(sr,sr+dict_soildata[k].__len__()):
              sheet.cell(sr-2,j+5,k)
              sheet.cell(sr-1,j+1,'layerId')
              sheet.cell(sr-1,j+2,'layerName')
              sheet.cell(sr-1,j+3,'qis')
              sheet.cell(sr-1,j+4,'len')
              sheet.cell(sr-1,j+5,'侧阻力i')
              sheet.cell(x,j+1,dict_soildata[k][x-sr]['layerId'])
              sheet.cell(x,j+2,dict_soildata[k][x-sr]['layerName'])
              sheet.cell(x,j+3,dict_soildata[k][x-sr]['qsi'])
              sheet.cell(x,j+4,dict_soildata[k][x-sr]['startHeight']-dict_soildata[k][x-sr]['endHeight'])
              sheet.cell(x,j+5,(dict_soildata[k][x-sr]['startHeight']-dict_soildata[k][x-sr]['endHeight'])*pile.cir*dict_soildata[k][x-sr]['qsi'])
    j=j+5

workbook.save('data.xlsx')


